"""
Journalisation des prédictions dans un fichier Excel.
"""
 
from __future__ import annotations 

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

ROOT_DIR = Path(__file__).resolve().parent.parent
DEFAULT_LOG_PATH = ROOT_DIR / "data" / "user_predictions.xlsx"

LOG_COLUMNS = [
    "timestamp",
    "username",
    "patient_id",
    "predicted_label",
    "success_probability",
    "risk_level",
    "warning_count",
    "warnings",
    "Recipientgender",
    "Stemcellsource",
    "Donorage",
    "DonorABO",
    "RecipientABO",
    "RecipientRh",
    "DonorCMV",
    "RecipientCMV",
    "Disease",
    "Riskgroup",
    "Txpostrelapse",
    "HLAmatch",
    "Antigen",
    "Alel",
    "HLAgrI",
    "Recipientage",
    "CD34kgx10d6",
    "CD3dCD34",
    "CD3dkgx10d8",
    "Rbodymass",
    "Gendermatch",
    "ABOmatch",
    "CMVstatus",
    "Diseasegroup",
    "HLAmismatch",
    "Recipientage10",
    "Recipientageint",
    "Donorage35",
    "clinical_note",
]


def _style_header_row(worksheet) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font


def initialize_workbook(path: str | Path = DEFAULT_LOG_PATH) -> Path:
    """
    Crée le fichier Excel s'il n'existe pas.
    """
    log_path = Path(path)
    log_path.parent.mkdir(parents=True, exist_ok=True)

    if log_path.exists():
        return log_path

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "predictions"
    sheet.append(LOG_COLUMNS)
    sheet.freeze_panes = "A2"
    _style_header_row(sheet)
    sheet.auto_filter.ref = f"A1:AK1"

    legend = workbook.create_sheet("legend")
    legend.append(["field", "meaning"])
    _style_header_row(legend)
    legend_rows = [
        ("predicted_label", "Survie ou non-survie prédite par le modèle"),
        ("success_probability", "Probabilité de succès clinique renvoyée par predict_proba"),
        ("risk_level", "Faible / moyen / élevé selon la probabilité"),
        ("Gendermatch", "1 = donneuse vers receveur homme ; 0 = autre combinaison"),
        ("ABOmatch", "1 = incompatibilité ABO ; 0 = groupe ABO identique"),
        ("CMVstatus", "0 = D-/R-, 1 = D+/R-, 2 = D-/R+, 3 = D+/R+"),
        ("Diseasegroup", "1 = maladie maligne ; 0 = non maligne"),
        ("HLAmismatch", "1 = HLA non parfait ; 0 = 10/10"),
        ("Recipientage10", "1 = receveur de 10 ans ou plus"),
        ("Recipientageint", "0 = <=5 ans ; 1 = ]5,10] ; 2 = >10"),
        ("Donorage35", "1 = donneur de 35 ans ou plus"),
    ]
    for row in legend_rows:
        legend.append(row)

    workbook.save(log_path)
    return log_path


def _risk_label(probability: float) -> str:
    if probability >= 0.70:
        return "Favorable"
    if probability >= 0.45:
        return "Intermédiaire"
    return "Élevé"


def append_prediction(
    username: str,
    raw_inputs: dict[str, Any],
    engineered_inputs: dict[str, Any],
    predicted_label: str,
    success_probability: float,
    warnings: list[str] | None = None,
    path: str | Path = DEFAULT_LOG_PATH,
) -> Path:
    """
    Ajoute une ligne dans le fichier Excel.
    """
    log_path = initialize_workbook(path)
    workbook = load_workbook(log_path)
    sheet = workbook["predictions"]

    warning_messages = warnings or []
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        username,
        raw_inputs.get("PatientID", ""),
        predicted_label,
        round(float(success_probability), 4),
        _risk_label(float(success_probability)),
        len(warning_messages),
        " | ".join(warning_messages),
        raw_inputs.get("Recipientgender"),
        raw_inputs.get("Stemcellsource"),
        raw_inputs.get("Donorage"),
        raw_inputs.get("DonorABO"),
        raw_inputs.get("RecipientABO"),
        raw_inputs.get("RecipientRh"),
        raw_inputs.get("DonorCMV"),
        raw_inputs.get("RecipientCMV"),
        raw_inputs.get("Disease"),
        raw_inputs.get("Riskgroup"),
        raw_inputs.get("Txpostrelapse"),
        raw_inputs.get("HLAmatch"),
        raw_inputs.get("Antigen"),
        raw_inputs.get("Alel"),
        raw_inputs.get("HLAgrI"),
        raw_inputs.get("Recipientage"),
        raw_inputs.get("CD34kgx10d6"),
        raw_inputs.get("CD3dCD34"),
        raw_inputs.get("CD3dkgx10d8"),
        raw_inputs.get("Rbodymass"),
        engineered_inputs.get("Gendermatch"),
        engineered_inputs.get("ABOmatch"),
        engineered_inputs.get("CMVstatus"),
        engineered_inputs.get("Diseasegroup"),
        engineered_inputs.get("HLAmismatch"),
        engineered_inputs.get("Recipientage10"),
        engineered_inputs.get("Recipientageint"),
        engineered_inputs.get("Donorage35"),
        raw_inputs.get("ClinicalNote", ""),
    ]

    sheet.append(row)
    workbook.save(log_path)
    return log_path


def read_history(path: str | Path = DEFAULT_LOG_PATH, limit: int = 50) -> pd.DataFrame:
    """
    Recharge l'historique pour affichage dans Streamlit.
    """
    log_path = initialize_workbook(path)
    df = pd.read_excel(log_path, sheet_name="predictions")
    if df.empty:
        return df
    return df.sort_values("timestamp", ascending=False).head(limit).reset_index(drop=True)
